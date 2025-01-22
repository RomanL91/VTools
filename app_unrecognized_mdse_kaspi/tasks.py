from celery import shared_task
from django.utils import timezone
from openpyxl import load_workbook

from .download import main
from .models import Product, ProductActivePeriod


@shared_task
def parse_pending_products():
    """
    1) Берём ВСЕ товары из БД (без фильтров).
    2) Читаем Excel-файл pending_products.xlsx (3 колонки: sku, name, manufacturer).
    3) Проходим по всем sku из файла, обновляем/создаём товары, при необходимости "реактивируем".
    4) Отдельно "закрываем" те товары, которых нет в файле, но которые сейчас in_file=True.
    """

    # -- 0. Загружаем Excel
    main()
    file_path = "pending_products.xlsx"  # <-- поменяйте на реальный путь
    wb = load_workbook(file_path)
    ws = wb.active  # предположим, нужные данные на первом листе

    # Собираем словарь sku -> (name, manufacturer) из файла,
    # чтобы потом легко проверять "есть ли sku в файле".
    file_data_map = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Предполагаем, что строки: (sku, name, manufacturer)
        sku, name, manufacturer = row
        if sku:
            # Если вдруг в файле случайно дубликат SKU, перезапишется последним
            file_data_map[sku] = (name, manufacturer)

    # Множество SKU, которые есть в файле (для удобного exclude)
    file_skus = set(file_data_map.keys())

    now_time = timezone.now()

    # -- 1. Берём все товары из БД
    all_products = Product.objects.all()  # без фильтра по in_file
    existing_products_map = {p.sku: p for p in all_products}

    # Для сборных операций
    new_products = []
    products_to_update = []
    new_periods = []

    # -- 2. Проход по всем SKU из файла
    for sku, (name, manufacturer) in file_data_map.items():
        if sku in existing_products_map:
            # Товар уже есть в БД
            product = existing_products_map[sku]
            needs_update = True

            # Если товар был "выключен" (in_file=False), а теперь опять в файле,
            # то "реактивируем" и открываем новый период
            if not product.in_file:
                product.in_file = True
                needs_update = True

                # Создаём новый период
                new_periods.append(
                    ProductActivePeriod(
                        product=product, started_at=now_time, ended_at=None
                    )
                )
            else:
                # in_file=True, но вдруг нужно проверить, закрыт ли старый период?
                # Если предыдущий период зачем-то закрыт, тоже открываем новый.
                # (Условие на ваше усмотрение.)
                has_open_period = product.active_periods.filter(
                    ended_at__isnull=True
                ).exists()
                if not has_open_period:
                    new_periods.append(
                        ProductActivePeriod(
                            product=product, started_at=now_time, ended_at=None
                        )
                    )

            if needs_update:
                product.updated_at = now_time
                products_to_update.append(product)

        else:
            # -- Новый товар
            new_p = Product(
                sku=sku,
                name=name,
                # manufacturer=manufacturer,
                in_file=True,
                created_at=now_time,
                updated_at=now_time,
            )
            new_products.append(new_p)

    # -- 3. Сохраняем новые товары
    created_products = Product.objects.bulk_create(new_products)

    # Создаём периоды для новых товаров
    for np in created_products:
        new_periods.append(
            ProductActivePeriod(product=np, started_at=now_time, ended_at=None)
        )

    # -- 4. bulk_update существующих товаров
    if products_to_update:
        Product.objects.bulk_update(
            products_to_update, fields=["name", "in_file", "updated_at"]
        )

    # -- 5. bulk_create новых периодов (как для вновь созданных, так и для "реактивированных")
    if new_periods:
        ProductActivePeriod.objects.bulk_create(new_periods)

    # -- 6. Второй проход:
    #      Найти товары, которые есть в БД (in_file=True),
    #      но НЕ в файле (sku не в file_skus).
    #      "Закрываем" их (ended_at=now), и ставим in_file=False.

    # Т.е. все товары, у которых sku не в file_skus и in_file=True
    # (значит до этого они считались "в списке"), но теперь пропали.
    missing_qs = Product.objects.filter(in_file=True).exclude(sku__in=file_skus)
    missing_products = list(missing_qs)
    if missing_products:
        # Закроем им периоды
        ProductActivePeriod.objects.filter(
            product__in=missing_products, ended_at__isnull=True
        ).update(ended_at=now_time)

        # И переключим in_file=False
        for mp in missing_products:
            mp.in_file = False
            mp.updated_at = now_time

        Product.objects.bulk_update(missing_products, fields=["in_file", "updated_at"])

    return (
        f"Processed  {len(file_skus)} SKU from file. "
        f"New products created: {len(new_products)}. "
        f"Closed (dropped from file): {len(missing_products)}."
    )
