import openpyxl
from openpyxl.utils import get_column_letter

from django.contrib import admin
from django.utils import timezone
from django.http import HttpResponse
from django.db.models import Sum, F, Case, When, ExpressionWrapper, DurationField

from .models import Product, ProductActivePeriod


class ProductActivePeriodInline(admin.TabularInline):
    model = ProductActivePeriod
    readonly_fields = ("started_at", "ended_at")
    extra = 0


@admin.register(Product)
class ProductAdmin(admin.ModelAdmin):
    list_display = [
        "name",
        "sku",
        "created_at",
        "updated_at",
        "total_active_time",
        "in_file",
    ]
    list_filter = [
        "in_file",
    ]
    search_fields = [
        "name",
        "sku",
    ]
    readonly_fields = [
        "name",
        "sku",
        "created_at",
        "updated_at",
        "total_active_time",
        "in_file",
    ]
    inlines = [ProductActivePeriodInline]
    actions = ["export_selected_to_excel"]

    def get_queryset(self, request):
        """Аннотируем QuerySet суммой длительностей всех периодов активности."""
        qs = super().get_queryset(request)

        now = timezone.now()

        # Для каждого active_period считаем:
        # если ended_at не заполнено -> считаем время (now - started_at)
        # иначе -> (ended_at - started_at).
        # Потом суммируем все периоды в одно поле `total_time`.
        total_active_expr = ExpressionWrapper(
            Case(
                When(
                    active_periods__ended_at__isnull=True,
                    then=now - F("active_periods__started_at"),
                ),
                default=F("active_periods__ended_at") - F("active_periods__started_at"),
                output_field=DurationField(),
            ),
            output_field=DurationField(),
        )

        qs = qs.annotate(total_time=Sum(total_active_expr))
        return qs

    def total_active_time(self, obj):
        """Вывод аннотированного поля, при желании можно красиво форматировать."""
        if obj.total_time is not None:
            # Пример форматирования: показать дни, часы, минуты
            total_seconds = obj.total_time.total_seconds()
            days = int(total_seconds // 86400)
            hours = int((total_seconds % 86400) // 3600)
            minutes = int((total_seconds % 3600) // 60)
            return f"{days}д {hours}ч {minutes}мин"
        return "-"

    total_active_time.short_description = "Суммарное время в списке"
    total_active_time.admin_order_field = "total_time"

    # Admin Action для экспорта в Excel
    def export_selected_to_excel(self, request, queryset):
        """
        Экспортирует выбранные товары в Excel-файл.
        """
        # Создаём рабочую книгу и лист
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Selected Products"

        # Определяем заголовки колонок
        headers = [
            "Артикул",
            "Название товара",
            "Нераспознанные KASPI",
            "Дата создания",
            "Дата обновления",
            "Суммарное время в списке",
        ]
        ws.append(headers)

        # Заполняем строки данными
        for product in queryset:
            row = [
                product.sku,
                product.name,
                "Да" if product.in_file else "Нет",
                product.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                product.updated_at.strftime("%Y-%m-%d %H:%M:%S"),
                product.total_active_time,
            ]
            ws.append(row)

        # Настраиваем ширину колонок (опционально)
        for i, column in enumerate(headers, 1):
            column_letter = get_column_letter(i)
            ws.column_dimensions[column_letter].width = max(len(column) + 2, 15)

        # Сохраняем рабочую книгу в буфер
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=selected_products.xlsx"
        wb.save(response)
        return response

    export_selected_to_excel.short_description = (
        "Экспортировать выбранные товары в Excel"
    )
